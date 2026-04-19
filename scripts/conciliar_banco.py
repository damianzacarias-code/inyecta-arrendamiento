#!/usr/bin/env python3
"""
Conciliación bancaria — Inyecta Arrendamiento (CLAUDE.md §9 T10)
================================================================

Lee uno o varios PDFs de estado de cuenta bancario, extrae las
transacciones (fecha, descripción, referencia, monto), y las cruza
contra los pagos (`payments`) y la amortización (`amortization_entries`)
de la base de datos del sistema. Produce un Excel con dos hojas:

    1. "Matches"      — transacciones con candidato asignado.
    2. "Pendientes"   — transacciones sin candidato confiable.

Reglas de matching (heurísticas, score 0-100):

    +40   Monto coincide con `pagoTotal` esperado del periodo (±$5)
    +35   Folio del contrato (ej "ARR-001-2026") aparece en la
          descripción o referencia
    +20   RFC del cliente aparece en la descripción o referencia
    +15   Fecha del depósito ≤ 5 días después del vencimiento
    +10   El periodo no tiene pago previo registrado

Score umbral: 50. Por debajo se considera "pendiente revisión manual".

Uso típico
----------
    cd sistema/
    python3 scripts/conciliar_banco.py \
        --pdf uploads/estados-cuenta/bbva_2026_03.pdf \
        --salida reportes/conciliacion_bbva_2026_03.xlsx

    # Procesar varios PDFs a la vez:
    python3 scripts/conciliar_banco.py \
        --pdf bbva.pdf santander.pdf banamex.pdf \
        --salida reportes/conciliacion_total.xlsx

    # Filtrar por banco y rango de fechas:
    python3 scripts/conciliar_banco.py --pdf bbva.pdf \
        --banco BBVA --desde 2026-03-01 --hasta 2026-03-31 \
        --salida reportes/marzo.xlsx

Variables de entorno
--------------------
    DATABASE_URL   tomado de server/.env si no está en el shell.
                   Formato: postgresql://user:pass@host:port/db

Notas
-----
- NO modifica la base de datos. Es un reporte de sólo lectura.
  Para confirmar matches, usa el endpoint POST /api/conciliation/match
  desde el frontend de cobranza (la lógica vive en
  server/src/routes/conciliation.ts).
- El parser de PDFs está calibrado para los layouts más comunes en MX
  (BBVA, Santander, Banamex, Banorte). Si encuentras un layout nuevo,
  agrega una regex al mapeo BANK_PATTERNS más abajo.
- Decimal nativo (no float) para todos los montos al comparar.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Optional

# ─── Dependencias externas (todas listadas en requirements.txt) ────
try:
    import pdfplumber
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from dotenv import load_dotenv
except ImportError as e:
    sys.stderr.write(
        f"Falta dependencia: {e.name}\n"
        f"Instalar con: pip install -r scripts/requirements.txt\n"
    )
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
SISTEMA_ROOT = SCRIPT_DIR.parent
SERVER_ENV = SISTEMA_ROOT / "server" / ".env"

# ════════════════════════════════════════════════════════════════════
# Tipos de datos
# ════════════════════════════════════════════════════════════════════


@dataclass
class TransaccionPDF:
    """Una línea extraída de un estado de cuenta PDF."""
    fecha: date
    descripcion: str
    referencia: Optional[str]
    monto: Decimal           # positivo = abono (depósito a Inyecta)
    tipo: str                # "ABONO" o "CARGO"
    pagina: int
    archivo: str             # nombre del PDF de origen


@dataclass
class PeriodoEsperado:
    """Una fila de amortización con su saldo pendiente actual."""
    contract_id: str
    contract_folio: str
    cliente_nombre: str
    cliente_rfc: Optional[str]
    periodo: int
    fecha_vencimiento: date
    pago_total: Decimal
    pagado_acumulado: Decimal

    @property
    def pendiente(self) -> Decimal:
        return (self.pago_total - self.pagado_acumulado).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    @property
    def descripcion_humana(self) -> str:
        return f"{self.contract_folio} · {self.cliente_nombre} · período {self.periodo}"


@dataclass
class Match:
    transaccion: TransaccionPDF
    candidato: PeriodoEsperado
    score: int
    razones: list[str] = field(default_factory=list)


# ════════════════════════════════════════════════════════════════════
# Parseo de PDFs bancarios
# ════════════════════════════════════════════════════════════════════

# Cada banco tiene un layout distinto. La estrategia es:
#   1) Detectar el banco por el header del PDF.
#   2) Aplicar un patrón de fila ad-hoc.
# El patrón debe capturar (fecha, descripcion, referencia?, monto, tipo).
# Para layouts no reconocidos hay un fallback genérico (BANK_GENERIC).

BANK_HEADER_PATTERNS = {
    "BBVA":      re.compile(r"\bBBVA\b", re.I),
    "SANTANDER": re.compile(r"\bSANTANDER\b", re.I),
    "BANAMEX":   re.compile(r"\bCITIBANAMEX|BANAMEX\b", re.I),
    "BANORTE":   re.compile(r"\bBANORTE\b", re.I),
    "HSBC":      re.compile(r"\bHSBC\b", re.I),
    "SCOTIA":    re.compile(r"\bSCOTIABANK|SCOTIA\b", re.I),
}

# Patrón genérico "DD/MM/YYYY  <texto>  <ref?>  $1,234.56"
# Acepta fecha en formatos DD/MM/YYYY, DD-MM-YYYY, DD MMM YYYY.
ROW_PATTERN_GENERIC = re.compile(
    r"""
    ^\s*
    (?P<fecha>\d{1,2}[\/\-\s][\w]{1,4}[\/\-\s]\d{2,4})   # fecha
    \s+
    (?P<descripcion>.+?)                                  # descripción (lazy)
    \s+
    (?P<monto>-?\$?\s?[\d,]+\.\d{2})                      # monto al final
    \s*$
    """,
    re.VERBOSE,
)

MESES_ES = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
    # Algunos PDFs usan abreviaturas en inglés
    "JAN": 1, "APR": 4, "AUG": 8, "DEC": 12,
}


def parsear_fecha(s: str, anio_default: int) -> Optional[date]:
    """Acepta DD/MM/YYYY, DD-MM-YYYY, DD MMM (con o sin año)."""
    s = s.strip().upper()
    # DD/MM/YYYY o DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        y = int(y)
        if y < 100:
            y += 2000
        try:
            return date(y, int(mo), int(d))
        except ValueError:
            return None
    # DD MMM YYYY o DD MMM
    m = re.match(r"^(\d{1,2})[\s\/\-]([A-Z]{3})[\s\/\-]?(\d{2,4})?$", s)
    if m:
        d, mo_txt, y = m.groups()
        if mo_txt not in MESES_ES:
            return None
        y = int(y) if y else anio_default
        if y < 100:
            y += 2000
        try:
            return date(y, MESES_ES[mo_txt], int(d))
        except ValueError:
            return None
    return None


def parsear_monto(s: str) -> Decimal:
    """Convierte '$1,234.56' o '-1234.56' a Decimal."""
    s = s.replace("$", "").replace(",", "").replace(" ", "").strip()
    return Decimal(s)


def detectar_banco(texto: str) -> str:
    for banco, patron in BANK_HEADER_PATTERNS.items():
        if patron.search(texto):
            return banco
    return "DESCONOCIDO"


def extraer_transacciones(pdf_path: Path) -> tuple[str, list[TransaccionPDF]]:
    """Recorre el PDF y devuelve (banco, [transacciones])."""
    transacciones: list[TransaccionPDF] = []
    banco = "DESCONOCIDO"
    anio_default = datetime.now().year

    with pdfplumber.open(str(pdf_path)) as pdf:
        for n, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text() or ""
            if n == 1:
                banco = detectar_banco(texto)
                # Intentar extraer el año del periodo del estado
                m = re.search(r"\b(20\d{2})\b", texto)
                if m:
                    anio_default = int(m.group(1))

            for linea in texto.splitlines():
                m = ROW_PATTERN_GENERIC.match(linea)
                if not m:
                    continue
                fecha = parsear_fecha(m.group("fecha"), anio_default)
                if fecha is None:
                    continue
                try:
                    monto = parsear_monto(m.group("monto"))
                except Exception:
                    continue
                descripcion = m.group("descripcion").strip()
                # La referencia suele ser un token alfanumérico al final
                # de la descripción
                ref_match = re.search(r"\b([A-Z0-9]{6,})\b", descripcion)
                referencia = ref_match.group(1) if ref_match else None

                tipo = "ABONO" if monto > 0 else "CARGO"
                transacciones.append(
                    TransaccionPDF(
                        fecha=fecha,
                        descripcion=descripcion,
                        referencia=referencia,
                        monto=monto,
                        tipo=tipo,
                        pagina=n,
                        archivo=pdf_path.name,
                    )
                )

    return banco, transacciones


# ════════════════════════════════════════════════════════════════════
# Carga de periodos esperados desde la BD
# ════════════════════════════════════════════════════════════════════

SQL_PERIODOS_PENDIENTES = """
SELECT
    c.id            AS contract_id,
    c.folio         AS contract_folio,
    COALESCE(cl."razonSocial",
             TRIM(CONCAT(cl.nombre, ' ', cl."apellidoPaterno"))) AS cliente_nombre,
    cl.rfc          AS cliente_rfc,
    a.periodo       AS periodo,
    a."fechaPago"   AS fecha_vencimiento,
    a."pagoTotal"   AS pago_total,
    COALESCE((
        SELECT SUM(p."montoTotal")
        FROM payments p
        WHERE p."contractId" = a."contractId"
          AND p.periodo = a.periodo
    ), 0)           AS pagado_acumulado
FROM amortization_entries a
JOIN contracts c   ON c.id = a."contractId"
JOIN clients   cl  ON cl.id = c."clientId"
WHERE c.estatus IN ('VIGENTE', 'VENCIDO')
"""


def cargar_periodos(
    conn,
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
) -> list[PeriodoEsperado]:
    """Carga los periodos pendientes (con saldo > 0) opcionalmente filtrados."""
    sql = SQL_PERIODOS_PENDIENTES
    params: list[object] = []
    # Ventana ±15 días alrededor del rango pedido para capturar pagos atrasados
    if desde:
        sql += ' AND a."fechaPago" >= %s'
        params.append(desde - timedelta(days=15))
    if hasta:
        sql += ' AND a."fechaPago" <= %s'
        params.append(hasta + timedelta(days=15))

    out: list[PeriodoEsperado] = []
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql, params)
        for row in cur.fetchall():
            pe = PeriodoEsperado(
                contract_id      = row["contract_id"],
                contract_folio   = row["contract_folio"],
                cliente_nombre   = (row["cliente_nombre"] or "").strip(),
                cliente_rfc      = row["cliente_rfc"],
                periodo          = row["periodo"],
                fecha_vencimiento= row["fecha_vencimiento"]
                                   if isinstance(row["fecha_vencimiento"], date)
                                   else row["fecha_vencimiento"].date(),
                pago_total       = Decimal(row["pago_total"]),
                pagado_acumulado = Decimal(row["pagado_acumulado"]),
            )
            if pe.pendiente > Decimal("0.01"):
                out.append(pe)
    return out


# ════════════════════════════════════════════════════════════════════
# Algoritmo de matching
# ════════════════════════════════════════════════════════════════════

TOLERANCIA_MONTO = Decimal("5.00")    # ±$5 absoluto
TOLERANCIA_DIAS = 5                   # ±5 días vs vencimiento
SCORE_UMBRAL = 50                     # < 50 → pendiente revisión manual


def conciliar(
    transacciones: list[TransaccionPDF],
    periodos: list[PeriodoEsperado],
) -> tuple[list[Match], list[TransaccionPDF]]:
    """Devuelve (matches, pendientes_sin_candidato)."""
    matches: list[Match] = []
    pendientes: list[TransaccionPDF] = []

    for tx in transacciones:
        # Sólo consideramos abonos para cobranza
        if tx.tipo != "ABONO":
            continue

        contexto = (tx.descripcion + " " + (tx.referencia or "")).upper()
        mejor: Optional[Match] = None

        for pe in periodos:
            score = 0
            razones: list[str] = []

            # +40 monto coincide (con tolerancia)
            diff = abs(tx.monto - pe.pendiente)
            diff_total = abs(tx.monto - pe.pago_total)
            if min(diff, diff_total) <= TOLERANCIA_MONTO:
                score += 40
                razones.append(
                    f"monto {tx.monto} ≈ esperado {pe.pendiente}"
                )

            # +35 folio del contrato presente
            if pe.contract_folio and pe.contract_folio.upper() in contexto:
                score += 35
                razones.append(f"folio {pe.contract_folio} en descripción")

            # +20 RFC del cliente presente
            if pe.cliente_rfc and pe.cliente_rfc.upper() in contexto:
                score += 20
                razones.append("RFC del cliente en descripción")

            # +15 fecha próxima al vencimiento
            dias_dif = abs((tx.fecha - pe.fecha_vencimiento).days)
            if dias_dif <= TOLERANCIA_DIAS:
                score += 15
                razones.append(f"±{dias_dif}d del vencimiento")

            # +10 sin pago previo del periodo
            if pe.pagado_acumulado <= Decimal("0.01"):
                score += 10
                razones.append("período sin pago previo")

            if mejor is None or score > mejor.score:
                mejor = Match(transaccion=tx, candidato=pe, score=score, razones=razones)

        if mejor and mejor.score >= SCORE_UMBRAL:
            matches.append(mejor)
        else:
            pendientes.append(tx)

    return matches, pendientes


# ════════════════════════════════════════════════════════════════════
# Salida Excel
# ════════════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", fgColor="1B2A47")  # color "TOTAL" (CLAUDE.md §3)
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")


def escribir_excel(
    salida: Path,
    matches: list[Match],
    pendientes: list[TransaccionPDF],
    banco: str,
) -> None:
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ─── Hoja 1: Matches ───────────────────────────────────────────
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Matches"
    columnas = [
        ("Archivo",          15),
        ("Pág",               5),
        ("Fecha",            12),
        ("Descripción",      45),
        ("Referencia",       18),
        ("Monto",            14),
        ("Folio contrato",   18),
        ("Cliente",          30),
        ("RFC",              16),
        ("Período",           8),
        ("Pendiente esperado", 16),
        ("Score",             8),
        ("Razones del match", 60),
    ]
    for col_idx, (titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[celda.column_letter].width = ancho

    for fila, m in enumerate(matches, start=2):
        valores = [
            m.transaccion.archivo,
            m.transaccion.pagina,
            m.transaccion.fecha.strftime("%d-%m-%Y"),
            m.transaccion.descripcion[:200],
            m.transaccion.referencia or "",
            float(m.transaccion.monto),
            m.candidato.contract_folio,
            m.candidato.cliente_nombre,
            m.candidato.cliente_rfc or "",
            m.candidato.periodo,
            float(m.candidato.pendiente),
            m.score,
            " · ".join(m.razones),
        ]
        for col_idx, v in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col_idx, value=v)
            if fila % 2 == 0:
                c.fill = ALT_FILL
        ws.cell(row=fila, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=fila, column=11).number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"

    # ─── Hoja 2: Pendientes ────────────────────────────────────────
    ws2 = wb.create_sheet("Pendientes")
    columnas2 = [
        ("Archivo",     15),
        ("Pág",          5),
        ("Fecha",       12),
        ("Descripción", 60),
        ("Referencia",  20),
        ("Monto",       14),
        ("Tipo",        10),
    ]
    for col_idx, (titulo, ancho) in enumerate(columnas2, start=1):
        c = ws2.cell(row=1, column=col_idx, value=titulo)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[c.column_letter].width = ancho
    for fila, tx in enumerate(pendientes, start=2):
        valores = [
            tx.archivo,
            tx.pagina,
            tx.fecha.strftime("%d-%m-%Y"),
            tx.descripcion[:200],
            tx.referencia or "",
            float(tx.monto),
            tx.tipo,
        ]
        for col_idx, v in enumerate(valores, start=1):
            c = ws2.cell(row=fila, column=col_idx, value=v)
            if fila % 2 == 0:
                c.fill = ALT_FILL
        ws2.cell(row=fila, column=6).number_format = '"$"#,##0.00'
    ws2.freeze_panes = "A2"

    # ─── Hoja 3: Resumen ───────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen")
    resumen = [
        ("Banco detectado",            banco),
        ("Total transacciones leídas", len(matches) + len(pendientes)),
        ("Matches con score ≥ 50",     len(matches)),
        ("Pendientes manuales",        len(pendientes)),
        ("Score umbral",               SCORE_UMBRAL),
        ("Tolerancia monto (MXN)",     float(TOLERANCIA_MONTO)),
        ("Tolerancia días",            TOLERANCIA_DIAS),
        ("Generado",                   datetime.now().strftime("%d-%m-%Y %H:%M")),
    ]
    for fila, (k, v) in enumerate(resumen, start=1):
        c = ws3.cell(row=fila, column=1, value=k)
        c.font = Font(bold=True)
        ws3.cell(row=fila, column=2, value=v)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 28

    wb.save(str(salida))


# ════════════════════════════════════════════════════════════════════
# CLI
# ════════════════════════════════════════════════════════════════════


def conectar_db() -> "psycopg2.extensions.connection":
    if SERVER_ENV.exists():
        load_dotenv(dotenv_path=SERVER_ENV)
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.stderr.write(
            "Falta DATABASE_URL. Defínelo en server/.env o en el shell.\n"
        )
        sys.exit(2)
    return psycopg2.connect(url)


def parse_fecha_arg(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def main(argv: Optional[Iterable[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Conciliación bancaria por PDF — Inyecta Arrendamiento"
    )
    p.add_argument("--pdf", nargs="+", required=True,
                   help="Uno o más PDFs de estado de cuenta")
    p.add_argument("--salida", required=True,
                   help="Ruta del Excel de salida (.xlsx)")
    p.add_argument("--banco", default=None,
                   help="Forzar nombre de banco (override del autodetect)")
    p.add_argument("--desde", type=parse_fecha_arg, default=None,
                   help="Filtra periodos cuyo vencimiento ≥ esta fecha (YYYY-MM-DD)")
    p.add_argument("--hasta", type=parse_fecha_arg, default=None,
                   help="Filtra periodos cuyo vencimiento ≤ esta fecha (YYYY-MM-DD)")
    args = p.parse_args(list(argv) if argv is not None else None)

    pdfs = [Path(x) for x in args.pdf]
    for pdf in pdfs:
        if not pdf.exists():
            sys.stderr.write(f"No existe: {pdf}\n")
            return 1

    print(f"→ Leyendo {len(pdfs)} PDF(s)…")
    todas_tx: list[TransaccionPDF] = []
    bancos: set[str] = set()
    for pdf in pdfs:
        banco, txs = extraer_transacciones(pdf)
        bancos.add(banco)
        todas_tx.extend(txs)
        print(f"   {pdf.name}: {len(txs)} transacciones — banco detectado: {banco}")

    if not todas_tx:
        sys.stderr.write("No se extrajo ninguna transacción de los PDFs.\n")
        return 1

    banco_label = args.banco or (next(iter(bancos)) if len(bancos) == 1 else "MIXTO")

    print("→ Conectando a la base de datos…")
    conn = conectar_db()
    try:
        periodos = cargar_periodos(conn, desde=args.desde, hasta=args.hasta)
    finally:
        conn.close()
    print(f"   {len(periodos)} períodos pendientes en la ventana solicitada")

    print("→ Conciliando…")
    matches, pendientes = conciliar(todas_tx, periodos)
    print(f"   {len(matches)} matches  /  {len(pendientes)} pendientes manuales")

    salida = Path(args.salida)
    print(f"→ Escribiendo {salida}…")
    escribir_excel(salida, matches, pendientes, banco_label)
    print("Listo ✓")
    return 0


if __name__ == "__main__":
    sys.exit(main())
